import matplotlib.pyplot as plt
from pydicom import dcmread
from pydicom.data import get_testdata_file
from openpyxl import load_workbook
from pathlib import Path
import os
import shutil
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image

filePath= "Metadatos.xlsx"
meta = load_workbook(filePath)
sheet = meta.active
x = sheet.max_row

a=[] #lista para hacer luego numpy concatenate
for i in range(2,x+1):
    #contenido = os.listdir(sheet.cell(row=i, column= 17).value)

    with os.scandir("D:\\Asignaturas\\Phyton y JavaScript\\imágenes cancer\\manifest-1616439774456" + sheet.cell(row=i, column= 17).value) as ficheros: # me voy a la carpeta que me marca el excel y escaneo los elementos que hay en la carpeta y los llamo "ficheros"
        for fichero in ficheros: # hay varios ficheros, hago un bucle para que los vaya leyendo uno a uno, fichero va desde 0 hasta ficheros
           
            ds = dcmread(fichero) #para que lea la imagen
            arr = ds.pixel_array  #transformar la imagen a array
            #print(arr)
            
            if np.all(arr[:,1913]==0): #si la útima columna de la imagen es toda ceros (el pecho está a la izquierda)
                arr = arr[::-1,::-1]  #rotar la imagen para que el pecho esté a la derecha 
            
            else:
                arr = ds.pixel_array

            a.append(arr)
            


algo = np.concatenate(([a]), axis=1) #concatena las matrices para formar un tensor 
print(algo.shape) #hay x matrices, y cada una tiene 2294 filas y 1914 columnas 
z,x,y =np.nonzero(algo)
x1 = x.min()
x2 = x.max()
y1 = y.min()
y2 = y.max()
z1 = z.min()
z2 = z.max()
algo1= algo[z1:z2+1, x1:x2, y1:y2] #recortas el tensor 
print(algo1.shape) 

    