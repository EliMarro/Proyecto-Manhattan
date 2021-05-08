import matplotlib.pyplot as plt
from pydicom import dcmread
from pydicom.data import get_testdata_file
from openpyxl import load_workbook
from pathlib import Path
import os
import shutil
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image
from xlrd import open_workbook
import pandas as pd


# filePath= "CMMD_clinicaldata_revision.xlsx"
# meta = load_workbook(filePath)
# sheet = meta.active
# x = sheet.max_row

file_name = 'CMMD_clinicaldata_revision.xlsx'

xl_workbook = pd.ExcelFile(file_name)  # Load the excel workbook
df = xl_workbook.parse("Sheet1")  # Parse the sheet into a dataframe
aList = df['classification'].tolist()  # Cast the desired column into a python list
#print(aList)

arr = np.array(aList)
# arr1 = np.reshape(arr, (-1,2))
# print(arr1)

# if arr == 'Benign':
#     [0,1]

# else:
#     [1,0]

# for n, i in enumerate(arr1):
#     if i=='Benign':
#         arr1[n,2] = [1,0]
#     if i== 'Malignant':
#         arr[n,2] = [0,1]

# print(arr1)

for n, i in enumerate(arr):
    if i=='Benign':
        arr[n] = 0
    if i== 'Malignant':
        arr[n] = 1

print(arr)

