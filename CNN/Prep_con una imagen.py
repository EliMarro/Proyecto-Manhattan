import matplotlib.pyplot as plt
from pydicom import dcmread
from pydicom.data import get_testdata_file
from openpyxl import load_workbook
from pathlib import Path
import os
import shutil
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image
import pandas as pd
import numpy as np
import seaborn as sns
from sklearn.model_selection import train_test_split
from sklearn import datasets, metrics
from sklearn.preprocessing import LabelEncoder



filePath= "Metadatos.xlsx"
meta = load_workbook(filePath)
sheet = meta.active
x = sheet.max_row

a=[] #lista para hacer luego numpy concatenate
for i in range(2,402):
    #contenido = os.listdir(sheet.cell(row=i, column= 17).value)

    with os.scandir("E://BCS//manifest-1616439774456" + sheet.cell(row=i, column= 17).value) as ficheros: # me voy a la carpeta que me marca el excel y escaneo los elementos que hay en la carpeta y los llamo "ficheros"
        for fichero in ficheros: # hay varios ficheros, hago un bucle para que los vaya leyendo uno a uno, fichero va desde 0 hasta ficheros
            file_name = Path(fichero).stem # para cada fichero, me separas el nombre de la extension
            numero = file_name[2] # quiero el tercer elemento del nombre de la imagen (1-1/1-2/...) 


            if int(numero) == 1:
                ds = dcmread(fichero) #para que lea la imagen
                arr = ds.pixel_array  #transformar la imagen a array
                #print(arr)

                if np.all(arr[:,1913]==0): #si la útima columna de la imagen es toda ceros (el pecho está a la izquierda)
                    arr = arr[::-1,::-1]  #rotar la imagen para que el pecho esté a la derecha
                
                else:
                    arr = ds.pixel_array
                
                a.append(arr)
            
algo = np.concatenate(([a]), axis=1) #concatena las matrices para formar un tensor 
print(algo.shape) #hay x matrices, y cada una tiene 2294 filas y 1914 columnas 
z,x,y =np.nonzero(algo)
x1 = x.min()
x2 = x.max()
y1 = y.min()
y2 = y.max()
z1 = z.min()
z2 = z.max()
algo1= algo[z1:z2+1, x1:x2, y1:y2] #recortas el tensor 
print(algo1.shape)
     

# algo2 = np.divide(algo1,255.) #divide los valores de la matriz para que sean números entre cero y uno



X_train, X_test = train_test_split(algo1, test_size=0.2) #para dividir las imágenes en 80% train y 20% test
print(X_train.shape)
print(X_test.shape)


np.save('X_train',X_train)
train = np.load('X_train.npy')
print(X_train.shape)
#para comprobar que lo guarda bien

np.save('X_test',X_test)
test = np.load('X_test.npy')
print(X_test.shape)
#para comprobar que lo guarda bien



file = 'CMMD_con_1776_elementos (1).csv'

cancer = pd.read_csv(file)


#print(cancer)
#print (cancer.head(10))

# contamos el numero de cancerigenos y de benignos en la columna classification
#print(cancer['classification'].value_counts()) 

# para saber el tipo de dato que almacena cada columna
#print(cancer.dtypes)

# Ahora voy a transformar, de la columna "classification" (columna 5 empezando en 0) Benign en 0 y Malign en 1
encoder = LabelEncoder()
#cancer.iloc[:, 1] = encoder.fit_transform(cancer.iloc[:,1].values) 
encoder = encoder.fit_transform(cancer.iloc[:,5].values) 
print(encoder)
#print(cancer.shape)

n = 400
N = int(n*0.8)
arr = np.array(encoder)
print(arr.shape)
arr1 = arr[0:N]
arr2 = arr[N:n]
print(arr1.shape)

np.save('Y_test_orig',arr2)
Y_test = np.load('Y_test_orig.npy')
print(Y_test.shape)
# # para comprobar que lo guarda bien

np.save('Y_train',arr1)
Y_train = np.load('Y_train.npy')
print(Y_train.shape)
# para comprobar que lo guarda bien
            
            


