import matplotlib.pyplot as plt
from pydicom import dcmread
from pydicom.data import get_testdata_file
from openpyxl import load_workbook
from pathlib import Path
import os
import shutil
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from PIL import Image

from sklearn.model_selection import train_test_split
from sklearn import datasets, metrics
from sklearn.preprocessing import LabelEncoder

from itertools import chain



filePath= "Metadatos.xlsx"
meta = load_workbook(filePath)
sheet = meta.active
x = sheet.max_row

new_range = chain(range(2,11), range(461, 462)) #para meter la imagen más grande dataset


a=[] #lista para hacer luego numpy concatenate
for i in new_range:
    #contenido = os.listdir(sheet.cell(row=i, column= 17).value)

    with os.scandir("D:\\Asignaturas\\Phyton y JavaScript\\imágenes cancer\\manifest-1616439774456" + sheet.cell(row=i, column= 17).value) as ficheros: # me voy a la carpeta que me marca el excel y escaneo los elementos que hay en la carpeta y los llamo "ficheros"
        for fichero in ficheros: # hay varios ficheros, hago un bucle para que los vaya leyendo uno a uno, fichero va desde 0 hasta ficheros
            file_name = Path(fichero).stem # para cada fichero, me separas el nombre de la extension
            numero = file_name[2] # quiero el tercer elemento del nombre de la imagen (1-1/1-2/...) 


            if int(numero) == 1:
                ds = dcmread(fichero) #para que lea la imagen
                arr = ds.pixel_array  #transformar la imagen a array
                #print(arr)

                if np.all(arr[:,1913]==0): #si la útima columna de la imagen es toda ceros (el pecho está a la izquierda)
                    arr = arr[::-1,::-1]  #rotar la imagen para que el pecho esté a la derecha
                
                else:
                    arr = ds.pixel_array
                
                a.append(arr)
            
algo = np.concatenate(([a]), axis=1) #concatena las matrices para formar un tensor 
print(algo.shape) #hay x matrices, y cada una tiene 2294 filas y 1914 columnas 
z,x,y =np.nonzero(algo)
y1 = y.min()
y2 = y.max()
z1 = z.min()
z2 = z.max()
algo1= algo[z1:z2+1, :, y1:y2] #recortas el tensor
print(algo1.shape)


#algo2 = np.divide(algo1,255.) #divide los valores de la matriz para que sean números entre cero y uno

#para conseguir un menor tamaño para dárselo a la red

a0 = algo1[:, ::4, ::3]
a1 = algo1[:, 1::4, ::3]
a2 = algo1[:, ::4, 1::3]
a3 = algo1[:, 1::4, 1::3]

a_ = 0.25*(a0 +a1+ a2+ a3)
print(a_.shape)



X_train, X_test = train_test_split(a_, test_size=0.2) #para dividir las imágenes en 80% train y 20% test
print(X_train.shape)
print(X_test.shape)


np.save('X_train',X_train)
train = np.load('X_train.npy')
print(X_train.shape)
#para comprobar que lo guarda bien

np.save('X_test',X_test)
test = np.load('X_test.npy')
print(X_test.shape)
#para comprobar que lo guarda bien




            


