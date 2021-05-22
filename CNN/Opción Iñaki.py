import matplotlib.pyplot as plt
from numpy.core.fromnumeric import diagonal
from numpy.lib.function_base import kaiser
from pydicom import dcmread
from pydicom.data import get_testdata_file
from openpyxl import load_workbook
from pathlib import Path
import os
import shutil
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image
import pandas as pd
import numpy as np
import seaborn as sns
from sklearn.model_selection import train_test_split
from sklearn import datasets, metrics
from sklearn.preprocessing import LabelEncoder



filePath= "Metadatos.xlsx"
meta = load_workbook(filePath)
sheet = meta.active
x = sheet.max_row

a=[] #lista para hacer luego numpy concatenate
b=[]
for i in range(15, 30):
    #contenido = os.listdir(sheet.cell(row=i, column= 17).value)

    with os.scandir("C:\\Users\\Usuario\\Desktop\\manifest-1616439774456" + sheet.cell(row=i, column= 17).value) as ficheros: # me voy a la carpeta que me marca el excel y escaneo los elementos que hay en la carpeta y los llamo "ficheros"
        for fichero in ficheros: # hay varios ficheros, hago un bucle para que los vaya leyendo uno a uno, fichero va desde 0 hasta ficheros
            file_name = Path(fichero).stem # para cada fichero, me separas el nombre de la extension
            numero = file_name[2] # quiero el tercer elemento del nombre de la imagen (1-1/1-2/...) 


            if int(numero) == 1:
                ds = dcmread(fichero) #para que lea la imagen
                arr = ds.pixel_array  #transformar la imagen a array
                #print(arr)
                
                if np.all(arr[:,1913]==0): #si la útima columna de la imagen es toda ceros (el pecho está a la izquierda)
                    arr = arr[::-1,::-1]  #rotar la imagen para que el pecho esté a la derecha
                    
                else:
                    arr = ds.pixel_array

                    a.append(arr)
        
        idx = np.argwhere(np.all(arr == 0, axis=0))
        # print(idx.shape)
        diana = np.delete(arr, idx, axis=1)
        # print(diana.shape)       
        # idx2=np.argwhere(np.all(diana==0, axis=1))
        # diana2 = np.delete(diana,idx2,axis=0)
        # print(diana2.shape)
        
        b.append(diana.size)
        # print(b)
        


        
max_value = max(b)  

max_index = b.index(max_value)
print(max_index+1)  

print(max_value) 

            

#concatena las matrices para formar un tensor 
# print(algo.shape) #hay x matrices, y cada una tiene 2294 filas y 1914 columnas 



