import matplotlib.pyplot as plt
from pydicom import dcmread
from pydicom.data import get_testdata_file
from openpyxl import load_workbook
from pathlib import Path
import os
import shutil
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image

from sklearn.model_selection import train_test_split
from PIL import Image
#import tensorflow as tf 
# import torch





filePath= "Metadatos.xlsx"
meta = load_workbook(filePath)
sheet = meta.active
x = sheet.max_row

a=[] #lista para meter los tamaños de las imágenes 
for i in range(2,x+1):
    #contenido = os.listdir(sheet.cell(row=i, column= 17).value)

    with os.scandir(r"E:\BCS\manifest-1616439774456" + sheet.cell(row=i, column= 17).value) as ficheros: # me voy a la carpeta que me marca el excel y escaneo los elementos que hay en la carpeta y los llamo "ficheros"
        for fichero in ficheros: # hay varios ficheros, hago un bucle para que los vaya leyendo uno a uno, fichero va desde 0 hasta ficheros
            file_name = Path(fichero).stem # para cada fichero, me separas el nombre de la extension
            numero = file_name[2] # quiero el tercer elemento del nombre de la imagen (1-1/1-2/...) 


            if int(numero) == 1:
                ds = dcmread(fichero) #para que lea la imagen
                arr = ds.pixel_array  #transformar la imagen a array
                #print(arr)

                if np.all(arr[:,1913]==0): #si la útima columna de la imagen es toda ceros (el pecho está a la izquierda)
                    arr = arr[::-1,::-1]  #rotar la imagen para que el pecho esté a la derecha
                
                else:
                    arr = ds.pixel_array

        x,y =np.nonzero(arr)
        x1 = x.min()
        x2 = x.max()
        y1 = y.min()
        y2 = y.max()
        arr1= arr[x1:x2, y1:y2] #recortas las matrices
        #print(arr1.shape)
        #print(arr1.size)
        a.append(arr1.size)

#print(a) #lista con el tamaño de las matrices 
max_index = a.index(max(a)) # imprime la posición de la imagen con el máximo tamaño 
print(max_index)  # empieza a contar desde cero

            
                
                
            


#print(algo1)
#print(np.amax(algo1))

#print(torch.max(algo1,1))


#np.where(algo1,[y.max()])

# vals,row_idx,col_idx = algo1.max(2)
# z3 = vals.argmax(0)
# print(algo1[z3, row_idx, col_idx])

#print(algo1[tf.math.argmax(algo1, axis=1)])

#print(tf.math.argmax(algo1, axis=1))
# algo2 = np.where(tf.math.argmax(algo1, axis=1))
# print(algo2)

# algo2 = np.where(algo1.argmax(axis=-1))
# print(algo2)

# algo2 = tf.argmax(algo1, axis=1)
# print(algo2)

#print (np.where(([algo1.max(axis=1)])))

# algo2 = algo1[z.max(),:,:]
# index = algo2.max()
# print(index)
# print(algo2)


