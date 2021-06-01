import matplotlib.pyplot as plt
from pydicom import dcmread
from pydicom.data import get_testdata_file
from openpyxl import load_workbook
from pathlib import Path
import os
import shutil
import numpy as np


filePath= "Metadatos.xlsx"
meta = load_workbook(filePath)
sheet = meta.active
x = sheet.max_row

a=[] #lista para meter los tamaños de las imágenes 
for i in range(2,x+1):
    #contenido = os.listdir(sheet.cell(row=i, column= 17).value)

    with os.scandir("D:\\Asignaturas\\Python y JavaScript\\imágenes cancer\\manifest-1616439774456" + sheet.cell(row=i, column= 17).value) as ficheros: # me voy a la carpeta que me marca el excel y escaneo los elementos que hay en la carpeta y los llamo "ficheros"
        for fichero in ficheros: # hay varios ficheros, hago un bucle para que los vaya leyendo uno a uno, fichero va desde 0 hasta ficheros
            file_name = Path(fichero).stem # para cada fichero, me separas el nombre de la extension
            numero = file_name[2] # quiero el tercer elemento del nombre de la imagen (1-1/1-2/...) 


            if int(numero) == 1:
                ds = dcmread(fichero) #para que lea la imagen
                arr = ds.pixel_array  #transformar la imagen a array
                #print(arr)

                if np.all(arr[:,1913]==0): #si la útima columna de la imagen es toda ceros (el pecho está a la izquierda)
                    arr = arr[::-1,::-1]  #rotar la imagen para que el pecho esté a la derecha
                
                else:
                    arr = ds.pixel_array

        x,y =np.nonzero(arr)
        x1 = x.min()
        x2 = x.max()
        y1 = y.min()
        y2 = y.max()
        arr1= arr[x1:x2, y1:y2] #recortas las matrices
        #print(arr1.shape)
        #print(arr1.size)
        a.append(arr1.size)

print(a) #lista con el tamaño de las matrices 
max_index = a.index(max(a)) # imprime la posición de la imagen con el máximo tamaño 
print(max_index)  # empieza a contar desde cero
  
             
                
            





